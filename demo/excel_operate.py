excel = """
excel操作使用openpyxl
作用：excel存测试数据

一、excel组成
1、excel文件 == 表格对象   一个表格对象有多个表单对象        小区大门
2、表单 == 表单对象        一个表单对象有多个单元格对象       单元门
3、单元格 == 单元格对象     单元格对象是excel的最小组成部分   你家的门

二、操作步骤
from openpyxl import load_workbook
#第一步创建表格对象
wb = load_workbook(filename="cases.xlsx")
#第二步获取表单对象
sheet_obj = wb["Sheet1"]
#第三步获取单元格对象
cell_obj = sheet_obj["B2"]
cell_obj2 = sheet_obj.cell(3,2)
print(cell_obj)
#第四步
print(cell_obj.value)
print(cell_obj2.value)
#第五步关闭excel连接
wb.close()

#获取所有的表单对象
from openpyxl import load_workbook
wb = load_workbook(filename='case_work')
for i in wb.sheetnames():
    print(wb[i])
wb.close()

三、行操作
1、result = sheet_obj.iter_rows()
min_row=None: 最小的行索引值（索引从1开始，必须是int类型，默认是1）
max_row=None: 最小的行索引值（索引从1开始，必须是int类型，默认是最大的行）
min_col=None: 最小的行索引值（索引从1开始，必须是int类型，默认是1）
max_col=None: 最小的行索引值（索引从1开始，必须是int类型，默认是最大的列）
values_only=False: True获取到values_only值，False拿到的是对象需要通过vaule属性获得对应的值
注意点：切片是闭区间，包含其实和结束索引对应的行的数据。

2、获取最大行
sheet_obj.max_row

3、获取最大列
sheet_obj.max_column
四、列操作
 resul = sheet_obj.iter_cols(min_col=1,max_col=3,values_only=True)
 wb = load_workbook(filename='cases.xlsx',read_only=True)
 read_only=False:可以直接获取列的值
 read_only=True：如果是True，列切片不可以直接读取值
 
 五、写excel文件【了解】
 注意：写Excel文件的时候要关闭Excel，否则会报错
 1、写单元格（覆盖）
 sheet_obj['A1'] = 'py52'
 sheet_obj.cell(row=1, column=2,value='test_value')
 
 2、追加写单元格（不会覆盖原有数据）
 在excel最后行加入数据
 可迭代对象的每个元素会占一个单元格，一个可迭代对象的数据占一行，不会换行
 接受不可迭代对象：list,tuple,dict,生成器
 sheet_obj.append([1,3,54,6])
test_dict={"A":"test1","B":"test2}【追加行不用写】

"""

# 线性脚本
from openpyxl import load_workbook

wb = load_workbook(filename="cases.xlsx", read_only=False)
sheet_obj = wb["sheet1"]
# [{"用例id:"1","用例名称":"登录成功"},{}]
result = list(sheet_obj.iter_rows(values_only=True))
print(result)
data_list = []
title_list = result[0]
value_list = result[1:]
for case in value_list:
    data_list.append(dict(list(zip(title_list, case))))
print(data_list)
wb.close()

# excel封装
from pprint import pprint
from openpyxl import load_workbook


class HandleExcel:
    def __init__(self, file_name, sheet_name):
        # 创建表对象
        self.wb = load_workbook(filename=file_name)
        # 创建sheet对象
        self.sheet_obj = self.wb[sheet_name]

    def get_cases(self):
        case_list = []
        cases = list(self.sheet_obj.iter_rows(values_only=True))
        title = cases[0]  # 获取表头
        values = cases[1:]  # 获取表数据
        for case in values:
            case_dict = dict(list(zip(title, case)))  # 将表数据和表头进行打包，变成dict
            case_list.append(case_dict)  # 将每一个测试用例加到list当中，集中返回
        self.close_file()  # 关闭文件
        return case_list

    def close_file(self):
        self.wb.close()


if __name__ == '__main__':
    cl = HandleExcel(file_name="cases.xlsx", sheet_name="sheet1")
    result = cl.get_cases()
    pprint(result)
